"""
THE CITADEL — Clan Intelligence Board
-------------------------------------
Live dashboard + deep Excel report for clan leadership.

DATA PULLED
  - Clan profile & full roster (donations given AND received, activity)
  - Current river race: fame, decks used, decks today, boat attacks
  - War log: last 8 wars — clan rank, fame, trophy change each week
  - Member war matrix: decks used per member per past war (who shows up)

RUN LOCALLY
  pip3 install requests openpyxl
  export CR_API_TOKEN=your_token_here
  python3 dashboard.py

DEPLOYED (Render)
  Uses $PORT automatically; token comes from the CR_API_TOKEN env var.
  API goes through proxy.royaleapi.dev — whitelist IP 45.79.218.79 on the key.
"""

import os
import io
import json
import webbrowser
import threading
import time
from datetime import datetime, timezone
from urllib.parse import quote
from http.server import HTTPServer, BaseHTTPRequestHandler

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Config --------------------------------------------------------
CLAN_TAG = "#JYPQRRCC"          # The Citadel
PORT = int(os.environ.get("PORT", 8000))
TOKEN = os.environ.get("CR_API_TOKEN", "PASTE_YOUR_TOKEN_HERE")
BASE = os.environ.get("CR_BASE", "https://proxy.royaleapi.dev/v1")
WARLOG_WEEKS = 8                # how many past wars to analyze

# ---- Logo palette --------------------------------------------------
SUMI  = "363B41"; RED = "D93A2B"; BAND = "F6F1E6"; LINE = "DDD6C6"

# ---- Data layer ----------------------------------------------------
_cache = {"t": 0, "data": None}

def _get(path, headers):
    r = requests.get(f"{BASE}{path}", headers=headers, timeout=20)
    return r

def fetch_all(force=False):
    """Clan + current war + war log, cached for 60s so refreshes are cheap."""
    now = time.time()
    if not force and _cache["data"] and now - _cache["t"] < 60:
        return _cache["data"]

    headers = {"Authorization": f"Bearer {TOKEN}", "Accept": "application/json"}
    tag = quote(CLAN_TAG.strip().upper())

    r = _get(f"/clans/{tag}", headers)
    if r.status_code != 200:
        return {"error": r.status_code}
    clan = r.json()

    # Current river race
    try:
        w = _get(f"/clans/{tag}/currentriverrace", headers)
        if w.status_code == 200:
            wr = w.json()
            me = wr.get("clan") or {}
            clan["_war"] = {
                "state": wr.get("state"),
                "fame": me.get("fame"),
                "periodPoints": me.get("periodPoints"),
                "participants": me.get("participants") or [],
                "standings": [
                    {"name": c.get("name"), "tag": c.get("tag"),
                     "fame": c.get("fame"), "points": c.get("periodPoints"),
                     "parts": [
                         {"name": p.get("name"), "tag": p.get("tag"),
                          "decks": p.get("decksUsed", 0),
                          "today": p.get("decksUsedToday", 0),
                          "fame": p.get("fame", 0)}
                         for p in (c.get("participants") or [])
                     ]}
                    for c in (wr.get("clans") or [])
                ],
            }
    except requests.RequestException:
        pass

    # War log (past wars)
    try:
        wl = _get(f"/clans/{tag}/riverracelog?limit={WARLOG_WEEKS}", headers)
        if wl.status_code == 200:
            items = wl.json().get("items", [])
            weeks, member_hist, clan_hist = [], {}, []
            for item in items:
                date = (item.get("createdDate") or "")[:8]
                if not date:
                    continue
                weeks.append(date)
                for st in item.get("standings", []):
                    c = st.get("clan", {})
                    if c.get("tag") != clan.get("tag"):
                        continue
                    clan_hist.append({
                        "date": date, "rank": st.get("rank"),
                        "trophyChange": st.get("trophyChange"),
                        "fame": c.get("fame"),
                    })
                    for p in c.get("participants") or []:
                        member_hist.setdefault(p.get("tag"), {})[date] = {
                            "decks": p.get("decksUsed", 0),
                            "fame": p.get("fame", 0),
                        }
            clan["_warlog"] = {"weeks": weeks, "members": member_hist,
                              "clan": clan_hist}
    except requests.RequestException:
        pass

    _cache["t"] = now
    _cache["data"] = clan
    return clan


def fetch_player(tag):
    """Full player profile + recent battles for the detail pop-up."""
    headers = {"Authorization": f"Bearer {TOKEN}", "Accept": "application/json"}
    t = quote(tag.strip().upper())
    r = _get(f"/players/{t}", headers)
    if r.status_code != 200:
        return {"error": r.status_code}
    p = r.json()
    try:
        b = _get(f"/players/{t}/battlelog", headers)
        if b.status_code == 200:
            battles = []
            for bt in b.json()[:10]:
                me = (bt.get("team") or [{}])[0]
                op = (bt.get("opponent") or [{}])[0]
                battles.append({
                    "type": bt.get("type", ""),
                    "myCrowns": me.get("crowns", 0),
                    "opCrowns": op.get("crowns", 0),
                    "opName": op.get("name", ""),
                })
            p["_battles"] = battles
    except requests.RequestException:
        pass
    # war-week history: cross-reference their clan's river race log
    try:
        ctag = (p.get("clan") or {}).get("tag")
        if ctag:
            wl = _get(f"/clans/{quote(ctag)}/riverracelog?limit=8", headers)
            if wl.status_code == 200:
                hist = []
                for item in wl.json().get("items", []):
                    date = (item.get("createdDate") or "")[:8]
                    for st in item.get("standings", []):
                        cc = st.get("clan", {})
                        if cc.get("tag") != ctag:
                            continue
                        for part in cc.get("participants") or []:
                            if part.get("tag") == p.get("tag"):
                                hist.append({"date": date,
                                             "decks": part.get("decksUsed", 0),
                                             "fame": part.get("fame", 0)})
                p["_warHistory"] = hist
    except requests.RequestException:
        pass
    return p


def fetch_opponents(clan):
    """For each rival clan in the current war: their history (last 8 wars).
    Their live rosters (decks/fame) are already in standings parts."""
    war = clan.get("_war") or {}
    opps = []
    headers = {"Authorization": f"Bearer {TOKEN}", "Accept": "application/json"}
    for s in war.get("standings", []):
        if s.get("tag") == clan.get("tag"):
            continue
        opp = dict(s)
        try:
            wl = _get(f"/clans/{quote(s['tag'])}/riverracelog?limit=8", headers)
            if wl.status_code == 200:
                hist = []
                for item in wl.json().get("items", []):
                    date = (item.get("createdDate") or "")[:8]
                    for st in item.get("standings", []):
                        cc = st.get("clan", {})
                        if cc.get("tag") == s.get("tag"):
                            hist.append({"date": date, "rank": st.get("rank"),
                                         "trophyChange": st.get("trophyChange"),
                                         "fame": cc.get("fame"),
                                         "decks": sum(pp.get("decksUsed", 0)
                                                      for pp in (cc.get("participants") or []))})
                opp["history"] = hist
        except requests.RequestException:
            opp["history"] = []
        opps.append(opp)
    return opps


def _project(hist, fame_now, decks_now):
    """Weighted projection from war history (newest first) + current pace.
    Transparent heuristic: recent wars weigh more; projected final fame =
    fame so far + (expected remaining decks x historical fame-per-deck)."""
    fam = [h.get("fame") or 0 for h in hist]
    dks = [h.get("decks") or 0 for h in hist]
    wts = list(range(len(hist), 0, -1))
    wavg_f = (sum(f * w for f, w in zip(fam, wts)) / sum(wts)) if hist else None
    per_deck = [f / d for f, d in zip(fam, dks) if d]
    fpd = sum(per_deck) / len(per_deck) if per_deck else None
    used = [d for d in dks if d]
    avg_d = sum(used) / len(used) if used else None
    if fpd and avg_d:
        remaining = max(avg_d - decks_now, 0)
        proj = fame_now + remaining * fpd
    elif wavg_f is not None:
        proj = max(fame_now, wavg_f)
    else:
        proj = fame_now
    return {"histAvg": wavg_f, "famePerDeck": fpd, "avgDecks": avg_d,
            "projected": round(proj)}


def compute_forecast(clan, opps):
    """Project final fame for every clan in the race + our expected lineup."""
    wl = clan.get("_warlog") or {}
    mh = wl.get("members", {})
    decks_by_week = {}
    for wk in mh.values():
        for d, v in wk.items():
            decks_by_week[d] = decks_by_week.get(d, 0) + (v.get("decks") or 0)
    ours_hist = [{"fame": e.get("fame") or 0,
                  "decks": decks_by_week.get(e["date"], 0)}
                 for e in wl.get("clan", [])]
    war = clan.get("_war") or {}
    our_f = war.get("fame") or 0
    our_d = sum(p.get("decksUsed", 0) for p in war.get("participants", []))
    clans = [dict(name=clan.get("name"), tag=clan.get("tag"), us=True,
                  fameNow=our_f, decksNow=our_d,
                  **_project(ours_hist, our_f, our_d))]
    for o in opps:
        hist = [{"fame": h.get("fame") or 0, "decks": h.get("decks") or 0}
                for h in o.get("history", [])]
        f_now = o.get("fame") or 0
        d_now = sum(p.get("decks", 0) for p in o.get("parts", []))
        clans.append(dict(name=o.get("name"), tag=o.get("tag"), us=False,
                          fameNow=f_now, decksNow=d_now,
                          **_project(hist, f_now, d_now)))
    tot = sum(c["projected"] for c in clans) or 1
    for c in clans:
        c["share"] = round(c["projected"] / tot * 100)
    clans.sort(key=lambda c: -c["projected"])

    # our projected lineup: expected decks x personal fame-per-deck
    lineup, weeks = [], wl.get("weeks", [])
    for m in clan.get("memberList", []):
        wk = mh.get(m.get("tag"), {})
        ds = [wk[w]["decks"] for w in weeks if w in wk]
        fs = [wk[w].get("fame", 0) for w in weeks if w in wk]
        if not ds:
            continue
        wts = list(range(len(ds), 0, -1))
        exp_d = sum(d * w for d, w in zip(ds, wts)) / sum(wts)
        pd = [f / d for f, d in zip(fs, ds) if d]
        fpd = sum(pd) / len(pd) if pd else 0
        lineup.append({"name": m.get("name"), "expDecks": round(exp_d, 1),
                       "fpd": round(fpd), "expFame": round(exp_d * fpd),
                       "part": round(sum(1 for d in ds if d) / len(ds) * 100)})
    lineup.sort(key=lambda x: -x["expFame"])
    return {"clans": clans, "lineup": lineup}


def _seen_dt(s):
    if not s:
        return None
    try:
        return datetime.strptime(s[:15], "%Y%m%dT%H%M%S").replace(tzinfo=timezone.utc)
    except Exception:
        return None


# ---- Excel report (6 sheets) ---------------------------------------
_H = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HF = PatternFill("solid", fgColor=SUMI)
_B = Border(bottom=Side(style="thin", color=LINE))
_CT = Alignment(horizontal="center", vertical="center")
_LT = Alignment(horizontal="left", vertical="center")

def _head(ws, headers, widths):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _H; c.fill = _HF
        c.alignment = _LT if ci == 2 else _CT
    ws.row_dimensions[1].height = 24
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

def _row(ws, ri, vals, red=(), num=()):
    band = PatternFill("solid", fgColor=BAND) if ri % 2 == 0 else None
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        is_red = ci in red
        c.font = Font(name="Arial", size=10, bold=is_red,
                      color=RED if is_red else SUMI)
        c.alignment = _LT if ci == 2 else _CT
        c.border = _B
        if band:
            c.fill = band
        if ci in num:
            c.number_format = "#,##0"

def build_workbook(clan):
    mem = clan.get("memberList", [])
    war = clan.get("_war") or {}
    parts = {p.get("tag"): p for p in war.get("participants", [])}
    in_war = bool(parts)
    wl = clan.get("_warlog") or {}
    weeks = wl.get("weeks", [])
    mh = wl.get("members", {})

    wb = Workbook()

    # 1) Summary
    s = wb.active; s.title = "Summary"
    s.sheet_view.showGridLines = False
    s.merge_cells("A1:B1")
    s["A1"] = f"THE CITADEL   {clan.get('tag','')}"
    s["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    s["A1"].fill = PatternFill("solid", fgColor=SUMI)
    s["A1"].alignment = _LT
    s.row_dimensions[1].height = 34
    weekly = sum(m.get("donations", 0) for m in mem)
    avgT = round(sum(m.get("trophies", 0) for m in mem) / (len(mem) or 1))
    inactive = sum(1 for m in mem if m.get("donations", 0) == 0)
    lab = Font(name="Arial", size=11, bold=True, color=SUMI)
    val = Font(name="Arial", size=11, color=SUMI)
    for i, (k, v) in enumerate([
        ("Clan name", clan.get("name", "")),
        ("Members", f"{clan.get('members', len(mem))}/50"),
        ("Clan score", clan.get("clanScore", 0)),
        ("War trophies", clan.get("clanWarTrophies", "-")),
        ("Required to join", clan.get("requiredTrophies", 0)),
        ("Weekly donations", weekly),
        ("Average trophies", avgT),
        ("Inactive (0 donations)", inactive),
        ("War status", war.get("state", "not in war")),
        ("Current war fame", war.get("fame", "-")),
        ("Report generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]):
        r = i + 3
        s[f"A{r}"] = k; s[f"A{r}"].font = lab
        c = s[f"B{r}"]; c.value = v; c.font = val
        if isinstance(v, (int, float)):
            c.number_format = "#,##0"
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 30

    # 2) Roster
    rs = wb.create_sheet("Roster")
    heads = ["Rank", "Name", "Tag", "Role", "Lvl", "Trophies",
             "Given/wk", "Received/wk", "Net", "Last seen"]
    _head(rs, heads, [7, 20, 13, 11, 6, 10, 10, 12, 9, 15])
    rl = {"leader": "Leader", "coLeader": "Co-leader",
          "elder": "Elder", "member": "Member"}
    for ri, m in enumerate(mem, start=2):
        g = m.get("donations", 0); rec = m.get("donationsReceived", 0)
        seen = _seen_dt(m.get("lastSeen"))
        _row(rs, ri,
             [m.get("clanRank"), m.get("name", ""), m.get("tag", ""),
              rl.get(m.get("role"), m.get("role", "")), m.get("expLevel"),
              m.get("trophies", 0), g, rec, g - rec,
              seen.strftime("%b %d, %H:%M") if seen else "-"],
             red=(7,) if g == 0 else (), num=(6, 7, 8, 9))

    # 3) Current war
    if in_war:
        cw = wb.create_sheet("Current War")
        _head(cw, ["#", "Name", "Decks used", "Decks today", "Fame",
                   "Boat attacks", "Repair pts"],
              [5, 20, 12, 12, 10, 13, 11])
        ordered = sorted(parts.values(),
                         key=lambda p: -(p.get("fame") or 0))
        for ri, p in enumerate(ordered, start=2):
            used = p.get("decksUsed", 0)
            _row(cw, ri, [ri - 1, p.get("name", ""), used,
                          p.get("decksUsedToday", 0), p.get("fame", 0),
                          p.get("boatAttacks", 0), p.get("repairPoints", 0)],
                 red=(3,) if used < 4 else (), num=(5, 7))

    # 4) War history (clan level)
    ch = wl.get("clan", [])
    if ch:
        wh = wb.create_sheet("War History")
        _head(wh, ["Week", "Rank", "Fame", "Trophy change"],
              [12, 8, 12, 14])
        for ri, w in enumerate(ch, start=2):
            d = w["date"]
            tc = w.get("trophyChange") or 0
            _row(wh, ri, [f"{d[4:6]}/{d[6:8]}/{d[0:4]}", w.get("rank"),
                          w.get("fame"), tc],
                 red=(4,) if tc < 0 else (), num=(3,))

    # 5) War matrix (member x week decks used) — the leadership sheet
    if weeks:
        wm = wb.create_sheet("War Matrix")
        heads = ["Rank", "Name"] + [f"{w[4:6]}/{w[6:8]}" for w in weeks] + \
                ["Total decks", "Missed wars", "Avg fame/war", "Fame/deck"]
        _head(wm, heads, [7, 20] + [8] * len(weeks) + [12, 12, 13, 10])
        for ri, m in enumerate(mem, start=2):
            row = [m.get("clanRank"), m.get("name", "")]
            tot, missed, fames = 0, 0, []
            for w in weeks:
                rec = mh.get(m.get("tag"), {}).get(w)
                if rec is None:
                    row.append("—")     # not in clan that week
                else:
                    d = rec["decks"]
                    row.append(d)
                    tot += d
                    fames.append(rec.get("fame", 0))
                    if d == 0:
                        missed += 1
            avg_f = round(sum(fames) / len(fames)) if fames else 0
            fpd = round(sum(fames) / tot) if tot else 0
            row += [tot, missed, avg_f, fpd]
            _row(wm, ri, row,
                 red=(len(row) - 2,) if missed > 0 else (), num=(len(row) - 1,))

    # 6) Enemy intel (war weeks only)
    opps = clan.get("_opponents") or []
    if opps:
        er = wb.create_sheet("Enemy Rosters")
        _head(er, ["Clan", "Player", "War decks", "Decks today", "Fame"],
              [20, 20, 11, 12, 10])
        ri = 2
        for o in opps:
            for p in sorted(o.get("parts", []),
                            key=lambda p: -(p.get("fame") or 0)):
                d = p.get("decks", 0)
                _row(er, ri, [o.get("name", ""), p.get("name", ""),
                              d, p.get("today", 0), p.get("fame", 0)],
                     red=(3,) if d == 0 else (), num=(5,))
                ri += 1

        eh = wb.create_sheet("Enemy History")
        _head(eh, ["Clan", "Week", "Rank", "Fame", "Trophy change",
                   "Wins (of 8)"],
              [20, 11, 7, 11, 13, 11])
        ri = 2
        for o in opps:
            hist = o.get("history", [])
            wins = sum(1 for h in hist if h.get("rank") == 1)
            for h in hist:
                d = h["date"]; tc = h.get("trophyChange") or 0
                _row(eh, ri, [o.get("name", ""),
                              f"{d[4:6]}/{d[6:8]}/{d[0:4]}",
                              h.get("rank"), h.get("fame"), tc, wins],
                     red=(5,) if tc < 0 else (), num=(4,))
                ri += 1

    # 6b) Forecast (transparent heuristic model)
    fc = clan.get("_forecast")
    if fc and fc.get("clans"):
        fs = wb.create_sheet("Forecast")
        _head(fs, ["Clan", "Hist avg fame", "Fame/deck", "Fame now",
                   "Decks now", "Projected final", "Share %"],
              [20, 13, 11, 11, 11, 14, 9])
        for ri, c in enumerate(fc["clans"], start=2):
            _row(fs, ri,
                 [c["name"],
                  round(c["histAvg"]) if c.get("histAvg") else "—",
                  round(c["famePerDeck"]) if c.get("famePerDeck") else "—",
                  c.get("fameNow", 0), c.get("decksNow", 0),
                  c["projected"], f'{c["share"]}%'],
                 red=(1,) if c.get("us") else (), num=(2, 4, 6))

    # 7) Donations leaderboard
    dl = wb.create_sheet("Donations")
    _head(dl, ["#", "Name", "Given", "Received", "Net"],
          [5, 20, 10, 11, 9])
    ordered = sorted(mem, key=lambda m: -(m.get("donations") or 0))
    for ri, m in enumerate(ordered, start=2):
        g = m.get("donations", 0); rec = m.get("donationsReceived", 0)
        _row(dl, ri, [ri - 1, m.get("name", ""), g, rec, g - rec],
             red=(3,) if g == 0 else (), num=(3, 4, 5))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- The page -------------------------------------------------------
PAGE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>The Citadel — Intelligence Board</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Stardos+Stencil:wght@400;700&family=Zen+Kaku+Gothic+New:wght@400;500;700&family=Space+Mono:wght@400;700&display=swap" rel="stylesheet">
<style>
:root{
  --paper:#efe9db; --panel:#fbf8f1; --band:#f2ece0;
  --sumi:#363b41; --soft:#5c636b; --ash:#8e959c;
  --red:#d93a2b; --red2:#b83124; --bronze:#9a7b4f;
  --green:#3f8f4f; --line:rgba(54,59,65,.14);
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Zen Kaku Gothic New',system-ui,sans-serif;background:var(--paper);
  color:var(--sumi);min-height:100vh;padding:34px 20px 90px;-webkit-font-smoothing:antialiased}
.wrap{max-width:1100px;margin:0 auto}
.hero{position:relative;text-align:center;padding:26px 0 30px;overflow:hidden}
.hero>*{position:relative;z-index:1}
.logo{width:150px;height:150px;border-radius:50%;object-fit:cover;
  border:3px solid var(--sumi);box-shadow:0 8px 30px rgba(54,59,65,.25);
  margin-bottom:14px;background:var(--paper)}
h1{font-family:'Stardos Stencil',serif;font-weight:700;
  font-size:clamp(40px,9vw,88px);letter-spacing:.02em;line-height:.92;
  text-shadow:0 2px 0 rgba(255,255,255,.35)}
.tag{font-family:'Space Mono';font-size:13px;letter-spacing:.22em;margin-top:14px;opacity:.7}
.desc{color:var(--soft);max-width:620px;margin:12px auto 0;font-size:15px;line-height:1.5}
.bar-top{display:flex;align-items:center;justify-content:center;gap:18px;margin-top:22px;flex-wrap:wrap}
.pulse{display:inline-flex;align-items:center;gap:8px;font-family:'Space Mono';
  font-size:11px;letter-spacing:.12em;color:var(--soft);text-transform:uppercase}
.dot{width:8px;height:8px;border-radius:50%;background:var(--red);animation:beat 2.4s infinite}
@keyframes beat{0%{box-shadow:0 0 0 0 rgba(217,58,43,.5)}70%{box-shadow:0 0 0 9px rgba(217,58,43,0)}100%{box-shadow:0 0 0 0 rgba(217,58,43,0)}}
.dl{display:inline-flex;align-items:center;gap:8px;text-decoration:none;background:var(--red);
  color:#fff;font-family:'Space Mono';font-weight:700;font-size:12px;letter-spacing:.08em;
  text-transform:uppercase;padding:11px 18px;transition:background .15s}
.dl:hover{background:var(--red2)}
.search{display:flex;gap:8px;max-width:560px;margin:22px auto 0}
.search input{flex:1;min-width:0;padding:11px 14px;border:2px solid var(--sumi);
  background:var(--panel);font-family:'Space Mono';font-size:13px;color:var(--sumi);outline:none}
.search input::placeholder{color:var(--ash)}
.sbtn{background:var(--sumi);color:#fff;border:none;cursor:pointer;
  font-family:'Space Mono';font-weight:700;font-size:11px;letter-spacing:.08em;
  text-transform:uppercase;padding:0 16px}
.sbtn.alt{background:var(--red)}
.sbtn:hover{opacity:.9}
.rule{height:2px;background:var(--sumi);margin:8px 0 0;opacity:.85}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(135px,1fr));
  gap:1px;background:var(--line);border:1px solid var(--line);margin:32px 0}
.stat{background:var(--panel);padding:18px 14px;text-align:center}
.stat .n{font-family:'Space Mono';font-weight:700;font-size:24px}
.stat.accent .n{color:var(--red)}
.stat .l{font-size:10px;letter-spacing:.12em;text-transform:uppercase;color:var(--ash);margin-top:6px}
.bh{display:flex;align-items:baseline;justify-content:space-between;margin:34px 0 12px;padding:0 4px;flex-wrap:wrap;gap:8px}
.bh h2{font-family:'Stardos Stencil';font-weight:700;font-size:23px;letter-spacing:.04em}
.bh .sub{font-family:'Space Mono';font-size:11px;color:var(--ash);letter-spacing:.1em}
.panel{border:1px solid var(--line);background:var(--panel)}
.row{display:grid;grid-template-columns:38px 1fr 140px 92px 92px 74px 74px;
  align-items:center;gap:12px;padding:12px 16px;border-bottom:1px solid var(--line)}
.row:last-child{border-bottom:0}
.row:nth-child(even){background:var(--band)}
.row:hover{background:#efe7d6}
.rank{font-family:'Space Mono';font-size:15px;color:var(--ash);text-align:center}
.rank.top{color:var(--red);font-weight:700}
.who{min-width:0}
.name{font-weight:500;font-size:15px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
  display:flex;align-items:center;gap:9px}
.shield{width:10px;height:13px;flex:0 0 auto;clip-path:polygon(0 0,100% 0,100% 62%,50% 100%,0 62%)}
.role{font-family:'Space Mono';font-size:10px;letter-spacing:.08em;text-transform:uppercase;
  color:var(--ash);margin-top:3px}
.away{color:var(--red);font-weight:700}
.trophies{display:flex;flex-direction:column;gap:5px}
.tnum{font-family:'Space Mono';font-size:13px}
.tbar{height:4px;background:rgba(54,59,65,.1);overflow:hidden}
.tbar i{display:block;height:100%;background:linear-gradient(90deg,var(--soft),var(--red))}
.cell{font-family:'Space Mono';font-size:13px;text-align:center}
.cell .cap{display:block;font-size:9px;letter-spacing:.08em;color:var(--ash);text-transform:uppercase;margin-top:2px}
.cell.bad{color:var(--red);font-weight:700}
.cell.good{color:var(--green)}
.wartbl{width:100%;border-collapse:collapse;font-family:'Space Mono';font-size:12px}
.wartbl th{background:var(--sumi);color:#fff;padding:9px 8px;font-weight:700;
  font-size:10px;letter-spacing:.08em;text-transform:uppercase}
.wartbl td{padding:9px 8px;border-bottom:1px solid var(--line);text-align:center}
.wartbl td:nth-child(2){text-align:left;font-family:'Zen Kaku Gothic New';font-size:13px}
.wartbl tr:nth-child(even) td{background:var(--band)}
.wartbl .bad{color:var(--red);font-weight:700}
.wartbl .good{color:var(--green);font-weight:700}
.us td{background:rgba(217,58,43,.09)!important;font-weight:700}
.foot{text-align:center;margin-top:30px;font-family:'Space Mono';font-size:11px;color:var(--ash);letter-spacing:.1em}
.err{background:rgba(217,58,43,.08);border:1px solid var(--red);color:var(--red2);
  padding:20px;text-align:center;font-family:'Space Mono';font-size:13px;margin-top:30px;line-height:1.6}
.row.click{cursor:pointer}
.overlay{position:fixed;inset:0;background:rgba(54,59,65,.55);display:none;
  align-items:flex-start;justify-content:center;padding:40px 16px;z-index:50;overflow-y:auto}
.overlay.open{display:flex}
.card{background:var(--paper);border:2px solid var(--sumi);max-width:560px;width:100%;
  box-shadow:0 20px 60px rgba(0,0,0,.35)}
.card-head{background:var(--sumi);color:#fff;padding:18px 22px;display:flex;
  justify-content:space-between;align-items:center}
.card-head h3{font-family:'Stardos Stencil';font-size:22px;letter-spacing:.03em}
.card-head .x{cursor:pointer;font-family:'Space Mono';font-size:18px;color:#fff;
  background:none;border:none;padding:4px 8px}
.card-body{padding:20px 22px}
.pgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));
  gap:1px;background:var(--line);border:1px solid var(--line);margin-bottom:18px}
.pstat{background:var(--panel);padding:12px 10px;text-align:center}
.pstat .n{font-family:'Space Mono';font-weight:700;font-size:18px}
.pstat .l{font-size:9px;letter-spacing:.1em;text-transform:uppercase;color:var(--ash);margin-top:4px}
.bl{font-family:'Space Mono';font-size:12px;width:100%;border-collapse:collapse}
.bl td{padding:7px 6px;border-bottom:1px solid var(--line)}
.bl .w{color:var(--green);font-weight:700}.bl .l2{color:var(--red);font-weight:700}
.loading{text-align:center;padding:30px;font-family:'Space Mono';font-size:12px;color:var(--ash)}
@media(max-width:720px){
  .row{grid-template-columns:30px 1fr 84px;gap:10px}
  .trophies,.cell{display:none}
  .wartbl th:nth-child(n+5),.wartbl td:nth-child(n+5){display:none}
}
</style></head><body><div class="wrap">
<div class="hero">
  <img class="logo" src="/logo.png" alt="" onerror="this.style.display='none'">
  <h1>THE CITADEL</h1>
  <div class="tag" id="ctag"></div>
  <p class="desc" id="cdesc"></p>
  <div class="bar-top">
    <span class="pulse"><span class="dot"></span><span id="updated">connecting…</span></span>
    <a class="dl" href="/download.xlsx">⤓ Full Excel Report</a>
  </div>
  <div class="search">
    <input id="q" placeholder="Player #TAG · clan name or #TAG"
      onkeydown="if(event.key==='Enter')doSearchEnter()">
    <button class="sbtn" onclick="doSearch('player')">Player</button>
    <button class="sbtn alt" onclick="doSearch('clan')">Clan</button>
  </div>
</div>
<div class="rule"></div>
<div id="body"></div>
<div class="foot">THE CITADEL · INTELLIGENCE BOARD · REFRESHES EVERY 60s · CLICK A MEMBER FOR THEIR RECORD</div>
</div>
<div class="overlay" id="ov" onclick="if(event.target===this)closeCard()">
  <div class="card">
    <div class="card-head"><h3 id="pname">Player</h3>
      <button class="x" onclick="closeCard()">✕</button></div>
    <div class="card-body" id="pbody"><div class="loading">Summoning record…</div></div>
  </div>
</div>
<script>
function closeCard(){document.getElementById('ov').classList.remove('open');}

function doSearchEnter(){
  const v=document.getElementById('q').value.trim();
  if(!v) return;
  // Enter: a #tag alone is ambiguous, so default tags to player; words to clan
  doSearch(v.startsWith('#')?'player':'clan');
}
function doSearch(mode){
  let v=document.getElementById('q').value.trim();
  if(!v) return;
  if(mode==='player'){
    if(!v.startsWith('#')) v='#'+v;
    openPlayer(encodeURIComponent(v.toUpperCase()), v.toUpperCase());
  }else{
    if(v.startsWith('#')){
      openScout(encodeURIComponent(v.toUpperCase()), v.toUpperCase());
    }else{
      searchClans(v);
    }
  }
}
async function searchClans(name){
  const ov=document.getElementById('ov');
  document.getElementById('pname').textContent='Clan search: '+name;
  document.getElementById('pbody').innerHTML='<div class="loading">Searching the realms…</div>';
  ov.classList.add('open');
  try{
    const r=await fetch('/api/searchclans?name='+encodeURIComponent(name));
    const d=await r.json();
    if(d.error){document.getElementById('pbody').innerHTML=
      '<div class="loading">'+d.error+'</div>';return;}
    const items=d.items||[];
    if(!items.length){document.getElementById('pbody').innerHTML=
      '<div class="loading">No clans found by that name.</div>';return;}
    let html='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">Tap a clan for its full scout report</div>';
    html+='<table class="bl"><tr style="font-weight:700"><td>Clan</td><td>Score</td><td>Members</td><td>Region</td></tr>';
    items.forEach(c=>{
      html+=`<tr class="pickclan" data-tag="${encodeURIComponent(c.tag)}" data-name="${(c.name||'').replace(/"/g,'&quot;')}" style="cursor:pointer">
        <td>${c.name}<br><span style="color:var(--ash);font-size:10px">${c.tag}</span></td>
        <td>${(c.score||0).toLocaleString()}</td>
        <td>${c.members??'—'}/50</td><td>${c.loc||''}</td></tr>`;
    });
    html+='</table>';
    document.getElementById('pbody').innerHTML=html;
  }catch(e){
    document.getElementById('pbody').innerHTML='<div class="loading">Connection error.</div>';
  }
}

async function openPlayer(tag,name){
  const ov=document.getElementById('ov');
  document.getElementById('pname').textContent=name;
  document.getElementById('pbody').innerHTML='<div class="loading">Summoning record…</div>';
  ov.classList.add('open');
  try{
    const r=await fetch('/api/player?tag='+tag);
    const p=await r.json();
    if(p.error){document.getElementById('pbody').innerHTML=
      '<div class="loading">Could not load ('+p.error+'). Check the tag.</div>';return;}
    document.getElementById('pname').textContent=p.name||name;
    const wins=p.wins||0, losses=p.losses||0;
    const wr=wins+losses?Math.round(wins/(wins+losses)*100):0;
    const stats=[
      ['Level',p.expLevel??'—'],['Trophies',(p.trophies||0).toLocaleString()],
      ['Best',(p.bestTrophies||0).toLocaleString()],['Wins',wins.toLocaleString()],
      ['Losses',losses.toLocaleString()],['Win rate',wr+'%'],
      ['3-crown wins',(p.threeCrownWins||0).toLocaleString()],
      ['Max challenge wins',p.challengeMaxWins??'—'],
      ['Total donations',(p.totalDonations||0).toLocaleString()],
      ['Battles',(p.battleCount||0).toLocaleString()],
    ];
    let html='<div class="pgrid">'+stats.map(s=>
      `<div class="pstat"><div class="n">${s[1]}</div><div class="l">${s[0]}</div></div>`).join('')+'</div>';
    if(p.currentDeck&&p.currentDeck.length){
      html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">Current deck</div>';
      html+='<div style="font-size:13px;line-height:1.7;margin-bottom:16px">'+
        p.currentDeck.map(c=>c.name).join(' · ')+'</div>';
    }
    if(p._warHistory&&p._warHistory.length){
      const H=p._warHistory, n=H.length;
      const td=H.reduce((s,h)=>s+(h.decks||0),0), tf=H.reduce((s,h)=>s+(h.fame||0),0);
      const played=H.filter(h=>h.decks>0).length;
      html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">War efficiency — recruiting metrics</div>';
      html+='<div class="pgrid">'+[
        ['Avg fame/war',Math.round(tf/n).toLocaleString()],
        ['Fame per deck',td?Math.round(tf/td):'—'],
        ['Avg decks/war',(td/n).toFixed(1)],
        ['Participation',Math.round(played/n*100)+'%'],
      ].map(s=>`<div class="pstat"><div class="n">${s[1]}</div><div class="l">${s[0]}</div></div>`).join('')+'</div>';
      const cn=(p.clan&&p.clan.name)?' (with '+p.clan.name+')':'';
      html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">Past war weeks'+cn+'</div>';
      html+='<table class="bl" style="margin-bottom:16px"><tr style="font-weight:700"><td>Week</td><td>Decks used</td><td>Fame</td></tr>';
      for(const h of p._warHistory){
        html+=`<tr><td>${h.date.slice(4,6)}/${h.date.slice(6,8)}</td>
          <td class="${h.decks===0?'l2':''}">${h.decks}</td>
          <td>${(h.fame||0).toLocaleString()}</td></tr>`;
      }
      html+='</table>';
    }
    if(p._battles&&p._battles.length){
      html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">Recent battles</div>';
      html+='<table class="bl">';
      for(const b of p._battles){
        const won=b.myCrowns>b.opCrowns;
        html+=`<tr><td class="${won?'w':'l2'}">${won?'WIN':'LOSS'}</td>
          <td>${b.myCrowns}–${b.opCrowns}</td><td>vs ${b.opName}</td></tr>`;
      }
      html+='</table>';
    }
    document.getElementById('pbody').innerHTML=html;
  }catch(e){
    document.getElementById('pbody').innerHTML='<div class="loading">Connection error.</div>';
  }
}

async function openScout(tag,name){
  const ov=document.getElementById('ov');
  document.getElementById('pname').textContent='⚔ '+name;
  document.getElementById('pbody').innerHTML='<div class="loading">Scouting the enemy…</div>';
  ov.classList.add('open');
  try{
    const r=await fetch('/api/scout?tag='+tag);
    const cl=await r.json();
    if(cl.error){document.getElementById('pbody').innerHTML=
      '<div class="loading">Could not scout ('+cl.error+').</div>';return;}
    const mem=cl.memberList||[];
    const avg=Math.round(mem.reduce((s,m)=>s+(m.trophies||0),0)/(mem.length||1));
    const weekly=mem.reduce((s,m)=>s+(m.donations||0),0);
    const stats=[
      ['Clan score',(cl.clanScore||0).toLocaleString()],
      ['War trophies',cl.clanWarTrophies??'—'],
      ['Members',(cl.members||mem.length)+'/50'],
      ['Avg trophies',avg.toLocaleString()],
      ['Weekly donations',weekly.toLocaleString()],
      ['To join',(cl.requiredTrophies||0).toLocaleString()],
    ];
    let html='<div class="pgrid">'+stats.map(s=>
      `<div class="pstat"><div class="n">${s[1]}</div><div class="l">${s[0]}</div></div>`).join('')+'</div>';

    // merge their war participation (already in standings data)
    const warParts={};
    const sd=(window._standings||{})[decodeURIComponent(tag)];
    if(sd) (sd.parts||[]).forEach(p=>warParts[p.tag]=p);

    html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin-bottom:6px">Their roster — ranked by trophies</div>';
    html+='<table class="bl"><tr style="font-weight:700"><td>#</td><td>Name</td><td>🏆</td><td>War decks</td><td>Fame</td></tr>';
    mem.slice(0,50).forEach((m,i)=>{
      const p=warParts[m.tag]||{};
      const decks=p.decks!==undefined?p.decks:'—';
      html+=`<tr><td>${i+1}</td><td>${m.name}</td>
        <td>${(m.trophies||0).toLocaleString()}</td>
        <td class="${decks===0?'l2':''}">${decks}</td>
        <td>${p.fame!==undefined?p.fame.toLocaleString():'—'}</td></tr>`;
    });
    html+='</table>';

    // their war history
    const hist=cl._history||[];
    if(hist.length){
      const wins=hist.filter(h=>h.rank===1).length;
      html+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin:16px 0 6px">Their last '+hist.length+' wars — '+wins+' victories</div>';
      html+='<table class="bl"><tr style="font-weight:700"><td>Week</td><td>Result</td><td>Fame</td><td>Trophies</td></tr>';
      for(const h of hist){
        const tc=h.trophyChange||0;
        html+=`<tr><td>${h.date.slice(4,6)}/${h.date.slice(6,8)}</td>
          <td class="${h.rank===1?'w':''}">${h.rank===1?'Victory':'#'+h.rank}</td>
          <td>${(h.fame||0).toLocaleString()}</td>
          <td class="${tc<0?'l2':'w'}">${tc>0?'+':''}${tc}</td></tr>`;
      }
      html+='</table>';
    }
    document.getElementById('pbody').innerHTML=html;
  }catch(e){
    document.getElementById('pbody').innerHTML='<div class="loading">Connection error.</div>';
  }
}

async function runForecast(){
  const fc=document.getElementById('fc');
  fc.innerHTML='<div class="loading">Consulting the oracle… gathering enemy histories</div>';
  try{
    const r=await fetch('/api/forecast'); const d=await r.json();
    if(d.error){fc.innerHTML='<div class="loading">Forecast failed ('+d.error+').</div>';return;}
    let h='<table class="wartbl"><tr><th>Clan</th><th>Hist avg</th><th>Fame now</th><th>Projected</th><th>Share</th></tr>';
    (d.clans||[]).forEach(c=>{
      h+=`<tr class="${c.us?'us':''}"><td>${c.name}</td>
        <td>${c.histAvg?Math.round(c.histAvg).toLocaleString():'—'}</td>
        <td>${(c.fameNow||0).toLocaleString()}</td>
        <td><b>${c.projected.toLocaleString()}</b></td><td>${c.share}%</td></tr>`;
    });
    h+='</table>';
    if(d.lineup&&d.lineup.length){
      h+='<div style="font-family:\'Space Mono\';font-size:10px;letter-spacing:.1em;color:var(--ash);text-transform:uppercase;margin:14px 10px 6px;text-align:left">Our projected contributors — exp decks × fame/deck · red = under 50% participation</div>';
      h+='<table class="bl" style="margin:0 10px 12px;width:calc(100% - 20px)">';
      d.lineup.slice(0,15).forEach((m,i)=>{
        h+=`<tr><td>${i+1}</td><td>${m.name}</td><td>${m.expDecks} decks</td>
          <td>${m.fpd}/deck</td><td><b>${m.expFame.toLocaleString()}</b></td>
          <td class="${m.part<50?'l2':'w'}">${m.part}%</td></tr>`;
      });
      h+='</table>';
    }
    fc.style.padding='0'; fc.style.textAlign='left'; fc.innerHTML=h;
  }catch(e){fc.innerHTML='<div class="loading">Connection error.</div>';}
}

document.addEventListener('click',e=>{
  const pc=e.target.closest('tr.pickclan');
  if(pc){openScout(pc.dataset.tag,pc.dataset.name);return;}
  const sc=e.target.closest('tr.scout');
  if(sc){openScout(sc.dataset.tag,sc.dataset.name);return;}
  const row=e.target.closest('.row.click');
  if(row) openPlayer(row.dataset.tag,row.dataset.name);
});
</script>
<script>
const ROLE={leader:{c:'var(--red)',label:'Leader'},coLeader:{c:'var(--sumi)',label:'Co-leader'},
  elder:{c:'var(--bronze)',label:'Elder'},member:{c:'var(--ash)',label:'Member'}};

function ago(s){
  if(!s) return {txt:'',days:0};
  const iso=s.replace(/(\d{4})(\d{2})(\d{2})T(\d{2})(\d{2})(\d{2}).*/,'$1-$2-$3T$4:$5:$6Z');
  const mins=(Date.now()-new Date(iso))/60000;
  if(isNaN(mins)) return {txt:'',days:0};
  if(mins<60) return {txt:Math.round(mins)+'m ago',days:0};
  if(mins<1440) return {txt:Math.round(mins/60)+'h ago',days:0};
  const d=Math.round(mins/1440);
  return {txt:d+'d ago',days:d};
}

function render(c){
  if(c.error){
    document.getElementById('body').innerHTML=
      '<div class="err">API returned '+c.error+'. If 403: make sure the key whitelists 45.79.218.79.</div>';
    return;
  }
  document.getElementById('ctag').textContent=c.tag;
  document.getElementById('cdesc').textContent=c.description||'';

  const mem=c.memberList||[];
  const maxT=Math.max(...mem.map(m=>m.trophies),1);
  const weekly=mem.reduce((s,m)=>s+(m.donations||0),0);
  const avgT=Math.round(mem.reduce((s,m)=>s+m.trophies,0)/(mem.length||1));
  const inactive=mem.filter(m=>(m.donations||0)===0).length;

  const war=c._war||{};
  const parts={}; (war.participants||[]).forEach(p=>parts[p.tag]=p);
  const inWar=(war.participants||[]).length>0;

  const wl=c._warlog||{}; const weeks=wl.weeks||[]; const mh=wl.members||{};

  // per-member war record across past weeks
  const record={};
  for(const m of mem){
    let tot=0,missed=0,played=0;
    for(const w of weeks){
      const r=(mh[m.tag]||{})[w];
      if(r!==undefined){played++;tot+=r.decks;if(r.decks===0)missed++;}
    }
    record[m.tag]={tot,missed,played};
  }

  // stat band
  const stats=[
    ['Clan score',(c.clanScore||0).toLocaleString(),1],
    ['War trophies',(c.clanWarTrophies!=null?c.clanWarTrophies:'—'),0],
    ['Members',(c.members||mem.length)+'/50',0],
    ['Weekly donations',weekly.toLocaleString(),0],
    ['Avg trophies',avgT.toLocaleString(),0],
    ['Inactive',inactive,0],
  ];
  if(inWar){
    stats.push(['War state',war.state||'—',1]);
    if(war.fame!=null) stats.push(['War fame',war.fame.toLocaleString(),1]);
  }
  let html='<div class="stats">'+stats.map(s=>
    `<div class="stat ${s[2]?'accent':''}"><div class="n">${s[1]}</div><div class="l">${s[0]}</div></div>`).join('')+'</div>';

  // current war standings
  if(inWar && (war.standings||[]).length){
    window._standings={};
    war.standings.forEach(s=>window._standings[s.tag]=s);
    html+=`<div class="bh"><h2>Current War</h2><span class="sub">RIVER RACE STANDINGS · CLICK A CLAN TO SCOUT</span></div>`;
    html+='<div class="panel" style="overflow-x:auto"><table class="wartbl"><tr><th>#</th><th>Clan</th><th>Fame</th><th>Points</th></tr>';
    const st=[...war.standings].sort((a,b)=>(b.fame||0)-(a.fame||0));
    st.forEach((s,i)=>{
      html+=`<tr class="${s.tag===c.tag?'us':''} scout" data-tag="${encodeURIComponent(s.tag)}" data-name="${(s.name||'').replace(/"/g,'&quot;')}" style="cursor:pointer">
        <td>${i+1}</td><td>${s.name||''}</td>
        <td>${(s.fame||0).toLocaleString()}</td><td>${(s.points||0).toLocaleString()}</td></tr>`;
    });
    html+='</table></div>';
  }

  // war forecast (always available; strongest mid-war)
  html+=`<div class="bh"><h2>War Forecast</h2>
    <span class="sub">WEIGHTED LAST-${weeks.length||8}-WAR MODEL · ESTIMATE, NOT PROPHECY</span></div>
  <div class="panel" id="fc" style="padding:18px;text-align:center">
    <button class="dl" onclick="runForecast()" style="border:none;cursor:pointer">⚡ Run Forecast</button>
  </div>`;

  // roster
  html+=`<div class="bh"><h2>The Roster</h2>
    <span class="sub">${inWar?'WAR WEEK · ':''}LAST ${weeks.length} WARS TRACKED · ${mem.length} STANDING</span></div>`;
  html+='<div class="panel">';
  for(const m of mem){
    const role=ROLE[m.role]||ROLE.member;
    const pct=Math.round(m.trophies/maxT*100);
    const g=m.donations||0, rec=record[m.tag]||{tot:0,missed:0,played:0};
    const seen=ago(m.lastSeen);
    const p=parts[m.tag]||{}; const used=p.decksUsed||0;
    const lvl=(m.expLevel&&m.expLevel>0)?` · lvl ${m.expLevel}`:'';
    html+=`<div class="row click" data-tag="${encodeURIComponent(m.tag)}" data-name="${m.name.replace(/"/g,'&quot;')}">
      <div class="rank ${m.clanRank<=3?'top':''}">${m.clanRank}</div>
      <div class="who">
        <div class="name"><span class="shield" style="background:${role.c}"></span>${m.name}</div>
        <div class="role">${role.label}${lvl} · <span class="${seen.days>=3?'away':''}">${seen.txt}</span></div>
      </div>
      <div class="trophies"><span class="tnum">🏆 ${m.trophies.toLocaleString()}</span>
        <span class="tbar"><i style="width:${pct}%"></i></span></div>
      <div class="cell ${g===0?'bad':''}">${g}<span class="cap">given/wk</span></div>
      ${inWar
        ? `<div class="cell ${used<4?'bad':'good'}">${used}<span class="cap">decks now</span></div>`
        : `<div class="cell">${m.donationsReceived||0}<span class="cap">received</span></div>`}
      <div class="cell">${rec.tot}<span class="cap">war decks</span></div>
      <div class="cell ${rec.missed>0?'bad':'good'}">${rec.missed}<span class="cap">missed</span></div>
    </div>`;
  }
  html+='</div>';

  // war history
  const ch=wl.clan||[];
  if(ch.length){
    html+=`<div class="bh"><h2>War History</h2><span class="sub">LAST ${ch.length} RIVER RACES</span></div>`;
    html+='<div class="panel" style="overflow-x:auto"><table class="wartbl"><tr><th>Week</th><th>Result</th><th>Rank</th><th>Fame</th><th>Trophies</th></tr>';
    for(const w of ch){
      const d=w.date, tc=w.trophyChange||0;
      html+=`<tr><td>${d.slice(4,6)}/${d.slice(6,8)}</td>
        <td>${w.rank===1?'Victory':'Placed #'+w.rank}</td><td>${w.rank??'—'}</td>
        <td>${(w.fame||0).toLocaleString()}</td>
        <td class="${tc<0?'bad':'good'}">${tc>0?'+':''}${tc}</td></tr>`;
    }
    html+='</table></div>';
  }

  document.getElementById('body').innerHTML=html;
}

async function load(){
  try{
    const r=await fetch('/api/clan');const c=await r.json();render(c);
    document.getElementById('updated').textContent='Live · updated '+new Date().toLocaleTimeString();
  }catch(e){document.getElementById('updated').textContent='connection lost — retrying';}
}
load();setInterval(load,60000);
</script></body></html>"""


# ---- Server ---------------------------------------------------------
class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype, extra=None):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        for k, v in (extra or {}).items():
            self.send_header(k, v)
        self.end_headers()
        if isinstance(body, str):
            body = body.encode("utf-8")
        self.wfile.write(body)

    def do_GET(self):
        if self.path == "/api/clan":
            self._send(200, json.dumps(fetch_all()), "application/json")
        elif self.path.startswith("/api/searchclans"):
            from urllib.parse import urlparse, parse_qs, unquote
            q = parse_qs(urlparse(self.path).query)
            name = unquote((q.get("name") or [""])[0]).strip()
            if len(name) < 3:
                self._send(200, '{"error":"Clan name must be at least 3 characters."}',
                           "application/json")
                return
            headers = {"Authorization": f"Bearer {TOKEN}",
                       "Accept": "application/json"}
            r = _get(f"/clans?name={quote(name)}&limit=10", headers)
            if r.status_code != 200:
                self._send(200, json.dumps({"error": r.status_code}),
                           "application/json")
                return
            items = [{"name": c.get("name"), "tag": c.get("tag"),
                      "score": c.get("clanScore"), "members": c.get("members"),
                      "loc": (c.get("location") or {}).get("name", "")}
                     for c in r.json().get("items", [])]
            self._send(200, json.dumps({"items": items}), "application/json")
        elif self.path.startswith("/api/scout"):
            from urllib.parse import urlparse, parse_qs, unquote
            q = parse_qs(urlparse(self.path).query)
            tag = unquote((q.get("tag") or [""])[0])
            if not tag.startswith("#"):
                self._send(400, '{"error":"bad tag"}', "application/json")
                return
            headers = {"Authorization": f"Bearer {TOKEN}",
                       "Accept": "application/json"}
            r = _get(f"/clans/{quote(tag)}", headers)
            body = r.json() if r.status_code == 200 else {"error": r.status_code}
            # their war history too
            if "error" not in body:
                try:
                    wl = _get(f"/clans/{quote(tag)}/riverracelog?limit=8", headers)
                    if wl.status_code == 200:
                        hist = []
                        for item in wl.json().get("items", []):
                            date = (item.get("createdDate") or "")[:8]
                            for st in item.get("standings", []):
                                cc = st.get("clan", {})
                                if cc.get("tag") == body.get("tag"):
                                    hist.append({
                                        "date": date,
                                        "rank": st.get("rank"),
                                        "trophyChange": st.get("trophyChange"),
                                        "fame": cc.get("fame"),
                                    })
                        body["_history"] = hist
                except requests.RequestException:
                    pass
            self._send(200, json.dumps(body), "application/json")
        elif self.path.startswith("/api/player"):
            from urllib.parse import urlparse, parse_qs, unquote
            q = parse_qs(urlparse(self.path).query)
            tag = unquote((q.get("tag") or [""])[0])
            if not tag.startswith("#"):
                self._send(400, '{"error":"bad tag"}', "application/json")
                return
            self._send(200, json.dumps(fetch_player(tag)), "application/json")
        elif self.path == "/api/forecast":
            clan = fetch_all()
            if clan.get("error"):
                self._send(200, json.dumps({"error": clan["error"]}),
                           "application/json")
                return
            self._send(200, json.dumps(
                compute_forecast(clan, fetch_opponents(clan))),
                "application/json")
        elif self.path == "/download.xlsx":
            clan = fetch_all()
            if clan.get("error"):
                self._send(502, "Could not reach the API.", "text/plain")
                return
            clan = dict(clan)
            clan["_opponents"] = fetch_opponents(clan)
            clan["_forecast"] = compute_forecast(clan, clan["_opponents"])
            data = build_workbook(clan)
            stamp = datetime.now().strftime("%Y-%m-%d")
            self._send(200, data,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       {"Content-Disposition":
                        f'attachment; filename="the-citadel-{stamp}.xlsx"'})
        elif self.path == "/logo.png":
            try:
                with open(os.path.join(os.path.dirname(
                        os.path.abspath(__file__)), "logo.png"), "rb") as f:
                    self._send(200, f.read(), "image/png",
                               {"Cache-Control": "public, max-age=86400"})
            except OSError:
                self._send(404, "no logo", "text/plain")
        elif self.path == "/":
            self._send(200, PAGE, "text/html; charset=utf-8")
        else:
            self._send(404, "not found", "text/plain")

    def log_message(self, *a):
        pass


def main():
    if TOKEN == "PASTE_YOUR_TOKEN_HERE":
        print("No token set. Run:  export CR_API_TOKEN=your_token_here")
        return
    local = os.environ.get("PORT") is None
    if local:
        url = f"http://localhost:{PORT}"
        print(f"The Citadel intelligence board: {url}")
        print("Press Control-C to stop.")
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    else:
        print(f"The Citadel board listening on port {PORT}")
    try:
        HTTPServer(("0.0.0.0", PORT), Handler).serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
